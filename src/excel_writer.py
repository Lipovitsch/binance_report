import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def adjust_path(path):
        if not path.endswith(".xlsx"):
            path += ".xlsx"
        return path


class ExcelWriter:
    def __init__(self, path):
        self.path: str = adjust_path(path)

    
    def save_dataframes_to_excel(self, dataframes: dict[str, pd.DataFrame], adjust = True):
        """
        Parameters
        ----------
        dataframes : dict[str, pd.DataFrame]
            dataframes dictionary should contain sheet name as a key and as a value dataframe, that should be saved inside this sheet
        
        Example
        -------
        See how the dataframes parameter should be built
        ```
        df1 = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        df2 = pd.DataFrame({"C": [1, 2], "D": [3, 4]})

        dataframes = {
            "Sheet_1": df1,
            "Sheet_2": df2
        }

        ew = ExcelWriter("example/path/excel.xlsx")
        ew.save_dataframes_to_excel(dataframes)
        ```
        """
        # Create new workbook
        workbook = Workbook()

        for sheet_name, data in dataframes.items():

            if type(data) == tuple:
                df: pd.DataFrame = data[0]
                export_index = data[1]
                # df = pd.DataFrame()
                if export_index:
                    df = df.reset_index()
            else:
                df = data


            # Create sheet for the current dataframe
            if len(dataframes) > 1:
                sheet = workbook.create_sheet(title=sheet_name)
            else:
                sheet = workbook['Sheet']
                sheet.title = sheet_name
            
            # Copy data from df to sheet
            sheet.append(df.columns.tolist())  # Add first row with column names
            for _, row in df.iterrows():
                sheet.append(row.tolist())

            def check_len(row):
                return max([len(el) for el in row.split('\n')])

            # Adjust columns width to the data
            if adjust:
                for col_num, column_name in enumerate(df.columns, 1):
                    max_length = max(
                        df[column_name].astype(str).map(check_len).max(),
                        len(str(column_name))
                    )
                    max_length = min(max_length, 80)
                    adjusted_width = (max_length) * 1.2
                    sheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

                    # Formatting columns as bold
                    sheet.cell(row=1, column=col_num).value = column_name
                    sheet.cell(row=1, column=col_num).font = Font(bold=True)

                    for row_num in range(2, len(df) + 2):  # Skip first row (col names)
                        cell = sheet.cell(row=row_num, column=col_num)
                        cell.alignment = Alignment(wrap_text=True)

        # Remove default sheet if there are more than one df to save
        if len(dataframes) > 1:
            workbook.remove(workbook["Sheet"])

        workbook.save(self.path)

